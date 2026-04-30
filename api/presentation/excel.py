from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from api.presentation.config import EXCEL_PATH

HEADERS = ["Fecha", "De", "Asunto", "Conversation ID", "Ruta HTML", "Adjuntos"]


def _get_workbook() -> Workbook:
    if EXCEL_PATH.exists():
        return load_workbook(EXCEL_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = "conversations"
    ws.append(HEADERS)
    return wb


def update_excel(record: dict) -> None:
    wb = _get_workbook()
    ws = wb.active

    # Find existing row by conversation_id (column 4, index 3)
    target_row: int | None = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=4).value == record["conversation_id"]:
            target_row = row_idx
            break

    values = [
        record["received_at"],
        record["from"],
        record["subject"],
        record["conversation_id"],
        record["html_path"],
        record["attachment_count"],
    ]

    if target_row:
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=target_row, column=col_idx, value=value)
    else:
        ws.append(values)

    wb.save(EXCEL_PATH)
